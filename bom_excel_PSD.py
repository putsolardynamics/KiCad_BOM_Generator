import kicad_netlist_reader
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Side, Border, Alignment
from openpyxl.utils import get_column_letter

# ======================================================================================================================
# Config
# ======================================================================================================================
vendors = ["Mouser", "Farnell", "TME", "Digikey"]
vendors_fields = ["Cena", "Koszt", "Link"]

# colors in aRGB format
header_color = "FFF7CB4D"
row_light = "FFFFFFFF"
row_dark = "FFFEF8E3"
border_color = "000000"
ending_color = "FFF7CB4D"

# ======================================================================================================================
# Create style objects
# ======================================================================================================================
thin = Side(border_style="thin", color=border_color)
medium = Side(border_style="medium", color=border_color)
border_default = Border(top=thin, left=thin, right=thin, bottom=thin)


# ======================================================================================================================
# File loading
# ======================================================================================================================
def my_equ(self, other):
    """myEqu is a more advanced equivalence function for components which is
    used by component grouping. Normal operation is to group components based
    on their value and footprint.

    In this example of a custom equivalency operator we compare the
    value, the part name and the footprint.
    """
    result = True
    if self.getValue() != other.getValue():
        result = False
    elif self.getPartName() != other.getPartName():
        result = False
    elif self.getFootprint() != other.getFootprint():
        result = False
    elif self.getDNP() != other.getDNP():
        result = False

    return result


# Override the component equivalence operator - it is important to do this
# before loading the netlist, otherwise all components will have the original
# equivalency operator.
kicad_netlist_reader.comp.__eq__ = my_equ

# Check if enough arguments were given
if len(sys.argv) != 3:
    print("Usage ", __file__, "<generic_netlist.xml> <output.csv>", file=sys.stderr)
    sys.exit(1)

# Generate an instance of a generic netlist, and load the netlist tree from
# the command line option. If the file doesn't exist, execution will stop
net = kicad_netlist_reader.netlist(sys.argv[1])
# subset the components to those wanted in the BOM, controlled
# by <configure> block in kicad_netlist_reader.py
components = net.getInterestingComponents(excludeBOM=True)

# Open spreadsheet and access first sheet
wb = Workbook()
ws = wb.active

# ======================================================================================================================
# Create rows
# ======================================================================================================================
# Create BOM header
# Add basic columns
columns = ['Id', 'Qty', 'Reference(s)', 'Value', 'Name', 'Footprint', 'Datasheet', 'DNP']
for idx, col in enumerate(columns):
    # convert column index to letter starting from A
    col_letter = get_column_letter(idx+1)
    # merge and add data
    ws.merge_cells(f'{col_letter}1:{col_letter}2')
    c = ws[f'{col_letter}1']
    c.value = col

# Add vendors columns
for idx, vendor in enumerate(vendors):
    # convert column index to letter starting from first free letter after basic columns
    vendor_letter = get_column_letter(1 + len(columns) + idx*len(vendors_fields))
    vendor_letter_2 = get_column_letter(1 + len(columns) + idx*len(vendors_fields) + len(vendors_fields)-1)
    # merge and add data
    ws.merge_cells(f'{vendor_letter}1:{vendor_letter_2}1')
    c = ws[f'{vendor_letter}1']
    c.value = vendor
    # add vendor fields cells
    for f_idx, field in enumerate(vendors_fields):
        field_letter = get_column_letter(1 + len(columns) + idx*len(vendors_fields) + f_idx)
        c = ws[f'{field_letter}2']
        c.value = field

# create data table
grouped = net.groupComponents(components)
components_table = list()
item_id = 0
row = list()
refs = ""
comp = None

for group in grouped:
    row = list()
    refs = ""
    # Add the reference of every component in the group and keep a reference
    # to the component so that the other data can be filled in once per group
    for component in group:
        if len(refs) > 0:
            refs += ", "
        refs += component.getRef()
        comp = component

    # Fill in the component groups common data
    item_id += 1
    row.append(item_id)
    row.append(len(group))
    row.append(refs)
    row.append(comp.getValue())
    row.append(comp.getLibName() + ":" + comp.getPartName())
    row.append(net.getGroupFootprint(group))
    row.append(net.getGroupDatasheet(group))
    row.append(comp.getDNPString())
    components_table.append(row)

# add data table to sheet
for row_idx, row in enumerate(components_table):
    for col_idx, value in enumerate(row):
        value_letter = get_column_letter(1 + col_idx)
        c = ws[f'{value_letter}{row_idx+3}']  # row_idx + 3 -> starting from letter C (A and B used by header)
        c.value = value

# add cost macros
for row_idx in range(len(components_table)):
    for col_idx in range(len(vendors)):
        col_letter = get_column_letter(1 + len(columns) + col_idx*len(vendors_fields) + 1)
        c = ws[f'{col_letter}{3 + row_idx}']
        col_letter = get_column_letter(1 + len(columns) + col_idx*len(vendors_fields))
        c.value = f"=B{3 + row_idx} * {col_letter}{3 + row_idx}"  # column B is hardcoded.
        # Change if general layout of columns is changed

# add cost sums
for col_idx in range(len(vendors)):
    col_letter = get_column_letter(1 + len(columns) + col_idx*len(vendors_fields))
    c = ws[f'{col_letter}{3 + len(components_table)}']
    c.value = "Suma:"
    col_letter = get_column_letter(1 + len(columns) + col_idx*len(vendors_fields) + 1)
    c = ws[f'{col_letter}{3 + len(components_table)}']
    c.value = f"=Sum({col_letter}3:{col_letter}{2 + len(components_table)})"

# ======================================================================================================================
# add style
# ======================================================================================================================
# header basic columns
for i in range(len(columns)):
    col_letter = get_column_letter(1 + i)
    c = ws[f'{col_letter}1']
    c.fill = PatternFill('solid', fgColor=header_color)
    c.border = border_default
    c.alignment = Alignment(horizontal='center', vertical='center')
    c = ws[f'{col_letter}2']
    c.border = Border(bottom=medium)

# header vendors
for i in range(len(vendors)):
    col_letter = get_column_letter(1 + len(columns) + i*len(vendors_fields))
    c = ws[f'{col_letter}1']
    c.fill = PatternFill('solid', fgColor=header_color)
    c.border = Border(top=thin, left=medium, right=thin, bottom=thin)
    c.alignment = Alignment(horizontal='center', vertical='center')
    # vendor fields
    for j in range(len(vendors_fields)):
        field_letter = get_column_letter(1 + len(columns) + i * len(vendors_fields) + j)
        c = ws[f'{field_letter}2']
        c.fill = PatternFill('solid', fgColor=header_color)
        c.alignment = Alignment(horizontal='center', vertical='center')
        if j == 0:
            c.border = Border(top=thin, left=medium, right=thin, bottom=medium)
        else:
            c.border = Border(top=thin, left=thin, right=thin, bottom=medium)

# data rows
light = True
for row_idx in range(len(components_table)):
    for col_idx in range(len(columns) + len(vendors)*len(vendors_fields)):
        col_letter = get_column_letter(1 + col_idx)
        c = ws[f'{col_letter}{3 + row_idx}']
        c.border = border_default
        c.fill = PatternFill('solid', fgColor=row_light) if light else PatternFill('solid', fgColor=row_dark)
    light = not light
    # add thicker line at every vendor beginning
    for col_idx in range(len(vendors)):
        col_letter = get_column_letter(1 + len(columns) + col_idx*len(vendors_fields))
        c = ws[f'{col_letter}{3 + row_idx}']
        c.border = Border(top=thin, left=medium, right=thin, bottom=thin)

# ending
for col_idx in range(len(columns) + len(vendors)*len(vendors_fields)):
    col_letter = get_column_letter(1 + col_idx)
    c = ws[f'{col_letter}{3 + len(components_table)}']
    c.fill = PatternFill('solid', fgColor=ending_color)

# adjust columns width
for idx, col in enumerate(columns, 1):
    ws.column_dimensions[get_column_letter(idx)].width = len(col) + 4

# write results
wb.save(sys.argv[2])
